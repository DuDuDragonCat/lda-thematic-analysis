# workers/fitlda_worker.py
# LDA 模型適配並輸出段落主題的後台邏輯（從 script/lda_fitAndGetBestPart_py.py 提取）

import os
import datetime
import math
import threading

import pandas as pd
import numpy as np
import tomotopy as tp
from openpyxl.worksheet.datavalidation import DataValidation

from workers.findbest_worker import DEFAULT as _FB_DEFAULT

RM_POS = ["V_2", "DE", "SHI", "FW", "I", "T", "WHITESPACE"]

DEFAULT = {
    # LDA 共享參數：以 findbest_worker.DEFAULT 為唯一來源，避免重複定義
    "seed": _FB_DEFAULT["seed"],
    "burnin": _FB_DEFAULT["burnin"],
    "iteration": _FB_DEFAULT["iteration"],
    "thin": _FB_DEFAULT["thin"],
    # Step4 私有參數
    "topicnum": "2",
    "words_num": "50",
    "exclude_single_char": _FB_DEFAULT["exclude_single_char"],
}


def safe_int(text, default):
    try:
        v = int(text)
        return v if v > 0 else default
    except (ValueError, TypeError):
        return default


def safe_bool(text, default=True):
    if isinstance(text, bool):
        return text
    if text is None:
        return default
    s = str(text).strip().lower()
    if s in {"1", "true", "yes", "y", "on"}:
        return True
    if s in {"0", "false", "no", "n", "off"}:
        return False
    return default


def make_filename(folder_path, topicnum, words_num, prefix, exclude_single_char):
    date = datetime.datetime.now().strftime("%Y%m%dT%H%M%S")
    single_char_tag = "rm1char" if exclude_single_char else "keep1char"
    filename = f"{prefix}_t{topicnum}_w{words_num}_{single_char_tag}.xlsx"
    filename_date = f"{prefix}_t{topicnum}_w{words_num}_{single_char_tag}_{date}.xlsx"
    if os.path.exists(os.path.join(folder_path, filename)):
        return os.path.join(folder_path, filename_date)
    return os.path.join(folder_path, filename)


def process_files(input_file_paths, part_words_num):
    """
    讀取斷詞後的 xlsx，進行詞性過濾，並將每份文件切成固定詞數的段落。
    """
    rawText_id_ary = []
    rawText_df_ary = []
    rawText_process_df_ary = []
    rawText_inputVecStr_ary = []
    partRawText_id_ary = []
    partRawText_df_ary = []
    partRawText_inputVecStr_ary = []
    partRawText_rawStr_ary = []

    for path in input_file_paths:
        filename = os.path.basename(path)
        rawText_id_ary.append(filename)

        df = pd.read_excel(path)
        df.reset_index(inplace=True)
        df.rename(columns={"index": "raw_index"}, inplace=True)
        rawText_df_ary.append(df)

        cols = ["raw_index", "sent_seq", "word", "pos", "sentsword_seq"]
        pre = df[cols].copy().dropna()
        pre = pre[~pre["pos"].str.endswith("CATEGORY")]
        pre = pre[~pre["pos"].isin(RM_POS)]
        pre.reset_index(drop=True, inplace=True)
        rawText_process_df_ary.append(pre)
        rawText_inputVecStr_ary.append(" ".join(pre["word"].values.tolist()))

        num_parts = math.ceil(pre.shape[0] / part_words_num)
        for i in range(num_parts):
            part_df = pre.iloc[i * part_words_num : (i + 1) * part_words_num]
            partRawText_df_ary.append(part_df)
            partRawText_inputVecStr_ary.append(
                " ".join(part_df["word"].values.tolist())
            )

            part_id = "{}_{}@{}-{}@{}".format(
                filename,
                part_df["sent_seq"].iloc[0],
                part_df["sentsword_seq"].iloc[0],
                part_df["sent_seq"].iloc[-1],
                part_df["sentsword_seq"].iloc[-1],
            )
            partRawText_id_ary.append(part_id)

            min_raw_index = int(part_df.iloc[0].raw_index)
            max_raw_index = int(part_df.iloc[-1].raw_index)
            tmp = df.copy()
            tmp["sent_seq"] = tmp["sent_seq"].astype(int)
            tmp["sentsword_seq"] = tmp["sentsword_seq"].astype(int)
            tmp["raw_index"] = tmp["raw_index"].astype(int)
            mask = (
                tmp.groupby("sent_seq")["sentsword_seq"].transform("max")
                == tmp["sentsword_seq"]
            )
            tmp.loc[mask, "word"] = tmp.loc[mask, "word"] + "\n"
            part_raw = tmp.query(
                "raw_index >= @min_raw_index & raw_index <= @max_raw_index"
            )["word"].values.tolist()
            partRawText_rawStr_ary.append("".join(part_raw))

    return (
        rawText_id_ary,
        rawText_df_ary,
        rawText_process_df_ary,
        rawText_inputVecStr_ary,
        partRawText_id_ary,
        partRawText_df_ary,
        partRawText_inputVecStr_ary,
        partRawText_rawStr_ary,
    )


def run_lda_tomotopy(
    raw_ids,
    raw_vec_strs,
    part_vec_strs,
    part_ids,
    k,
    seed,
    burn_in,
    iteration,
    thin,
    exclude_single_char,
):
    """
    兩階段 LDA：
      Stage1 — 用全文訓練模型（add_doc + train）。
      Stage2 — 對「段落」做 infer()（段落未參與訓練）；
               對「原始文章」直接讀 mdl.docs[i].get_topic_dist()
               （訓練時已充分取樣，無需重新 infer）。
    回傳 gamma_df（段落層級）、beta_df（詞-主題）、raw_gamma_df（文章層級）。
    """
    if exclude_single_char:
        raw_filtered = [[w for w in s.split() if len(w) >= 2] for s in raw_vec_strs]
        part_filtered = [[w for w in s.split() if len(w) >= 2] for s in part_vec_strs]
    else:
        raw_filtered = [s.split() for s in raw_vec_strs]
        part_filtered = [s.split() for s in part_vec_strs]

    mdl = tp.LDAModel(k=k, seed=seed)
    for doc in raw_filtered:
        if doc:
            mdl.add_doc(doc)

    if burn_in > 0:
        mdl.train(burn_in, workers=1)

    post_iters = max(0, iteration)
    if thin > 0 and post_iters > 0:
        for _ in range(max(1, post_iters // thin)):
            mdl.train(thin, workers=1)

    # ── 詞-主題分布（WordTopic） ─────────────────────────────────────────
    vocab = list(mdl.vocabs)
    beta_rows = []
    for t_idx in range(k):
        word_dist = mdl.get_topic_word_dist(t_idx)
        for w_idx, prob in enumerate(word_dist):
            prob = float(prob) if prob > 0 else 1e-12
            beta_rows.append(
                {
                    "topic": t_idx + 1,
                    "word": vocab[w_idx],
                    "log_probability": float(np.log(prob)),
                    "probability": prob,
                }
            )
    beta_df = (
        pd.DataFrame(beta_rows)
        .sort_values(["topic", "probability", "word"], ascending=[True, False, True])
        .reset_index(drop=True)
    )

    # ── 原始文章層級（TopicDoc）─────────────────────────────────────────
    # 訓練時 add_doc 的文件已在 mdl.docs 中保有完整的 Gibbs sampling 分布，
    # 直接讀取即可，不需要重新 infer（更準確且無額外計算開銷）。
    # 注意：空文件被跳過未 add_doc，故用 iterator 對齊。
    trained_doc_iter = iter(mdl.docs)
    raw_gamma_rows = []
    for doc_id, doc_words in zip(raw_ids, raw_filtered):
        if doc_words:
            topic_probs = np.array(next(trained_doc_iter).get_topic_dist())
        else:
            topic_probs = np.ones(k) / k
        for t_idx, prob in enumerate(topic_probs):
            raw_gamma_rows.append(
                {"document": doc_id, "topic": t_idx + 1, "gamma": float(prob)}
            )
    raw_gamma_df = pd.DataFrame(raw_gamma_rows)

    # ── 段落層級推斷（TopicPart） ─────────────────────────────────────────
    gamma_rows = []
    for doc_id, part_words in zip(part_ids, part_filtered):
        if part_words:
            doc_inst = mdl.make_doc(part_words)
            inferred, _ = mdl.infer(doc_inst, iterations=500, workers=1)
            topic_probs = np.array(inferred)
        else:
            topic_probs = np.ones(k) / k
        for t_idx, prob in enumerate(topic_probs):
            gamma_rows.append(
                {"document": doc_id, "topic": t_idx + 1, "gamma": float(prob)}
            )
    gamma_df = pd.DataFrame(gamma_rows)

    return gamma_df, beta_df, raw_gamma_df


def run_fitlda(
    input_files,
    output_folder,
    k,
    seed,
    burn_in,
    iteration,
    thin,
    words_num,
    exclude_single_char,
    on_append,
    on_done,
    on_error,
):
    """啟動背景執行緒執行 LDA 適配。"""

    def _worker():
        try:
            on_append(
                f"SEED: {seed}; Burn In: {burn_in}; Iteration: {iteration}; Thin: {thin}"
            )
            on_append(f"Topic Number: {k}; 每段落詞數: {words_num}")
            on_append(
                f"排除單字詞：{'是' if exclude_single_char else '否'}"
            )
            (raw_ids, _, _, raw_vec_strs, part_ids, _, part_vec_strs, part_raw_strs) = (
                process_files(input_files, words_num)
            )

            on_append(f"文章數：{len(raw_ids)}；段落數：{len(part_ids)}。")
            on_append("LDA 模型適配中（可能需要較長時間）...")

            gamma_df, beta_df, raw_gamma_df = run_lda_tomotopy(
                raw_ids,
                raw_vec_strs,
                part_vec_strs,
                part_ids,
                k,
                seed,
                burn_in,
                iteration,
                thin,
                exclude_single_char,
            )

            # ── TopicPart：段落層級主題分布 ────────────────────────────────
            gamma_pivot = gamma_df.pivot(
                index="document", columns="topic", values="gamma"
            ).reset_index()
            gamma_pivot.columns.name = None
            topic_cols = [c for c in gamma_pivot.columns if c != "document"]
            gamma_pivot.columns = ["document"] + [f"topic_{c}" for c in topic_cols]

            id_to_raw = dict(zip(part_ids, part_raw_strs))
            gamma_pivot["raw_text"] = gamma_pivot["document"].map(id_to_raw)

            # ── TopicDoc：原始文章層級主題分布 ────────────────────────────
            raw_pivot = raw_gamma_df.pivot(
                index="document", columns="topic", values="gamma"
            ).reset_index()
            raw_pivot.columns.name = None
            raw_topic_cols = [c for c in raw_pivot.columns if c != "document"]
            raw_pivot.columns = ["document"] + [f"topic_{c}" for c in raw_topic_cols]

            # ── TopicPart / TopicDoc：長表（第二分頁）────────────────────────
            gamma_long = gamma_df.rename(columns={"gamma": "probability"})
            gamma_long["raw_text"] = gamma_long["document"].map(id_to_raw)
            gamma_long = gamma_long.sort_values(
                ["topic", "probability", "raw_text", "document"],
                ascending=[True, False, True, True],
            ).reset_index(drop=True)
            raw_long = (
                raw_gamma_df.rename(columns={"gamma": "probability"})
                .sort_values(["document", "topic"])
                .reset_index(drop=True)
            )

            # ── 存檔 ──────────────────────────────────────────────────────
            topic_part_path = make_filename(
                output_folder, k, words_num, "TopicPart", exclude_single_char
            )
            word_topic_path = make_filename(
                output_folder, k, words_num, "WordTopic", exclude_single_char
            )
            topic_doc_path = make_filename(
                output_folder, k, words_num, "TopicDoc", exclude_single_char
            )

            with pd.ExcelWriter(topic_part_path, engine="openpyxl") as writer:
                gamma_pivot.to_excel(writer, index=False, sheet_name="TopicPart")
                gamma_long.to_excel(writer, index=False, sheet_name="TopicPart_Long")

                topk_sheet = "TopK_Results"
                ws = writer.book.create_sheet(title=topk_sheet)

                ws["A1"] = "TopicPart TopK 查詢"
                ws["A3"] = "topic number"
                ws["A4"] = "top number"
                ws["A5"] = "max rows/topic"
                ws["C3"] = '=IF(AND($B$3>=1,$B$3<=$E$1),"topic_"&$B$3,"")'

                ws["B3"] = 1
                max_rows_per_topic = (
                    int(gamma_long.groupby("topic").size().max())
                    if not gamma_long.empty
                    else 1
                )
                ws["B4"] = min(10, max_rows_per_topic)
                ws["B5"] = max_rows_per_topic
                ws["E1"] = int(k)

                ws["A6"] = "rank"
                ws["B6"] = "document"
                ws["C6"] = "probability"
                ws["D6"] = "raw_text"

                # 將每個 topic 依 probability 由高到低排序，固定為等長區塊，
                # 讓 TopK 區只需用 INDEX + 偏移量公式即可切換 topic 與 top number。
                helper_rows = []
                for topic_i in range(1, k + 1):
                    topic_df = gamma_long[gamma_long["topic"] == topic_i].sort_values(
                        ["probability", "raw_text", "document"],
                        ascending=[False, True, True],
                    )
                    doc_vals = topic_df["document"].tolist()
                    prob_vals = topic_df["probability"].tolist()
                    raw_vals = topic_df["raw_text"].tolist()
                    for row_i in range(max_rows_per_topic):
                        if row_i < len(topic_df):
                            helper_rows.append(
                                (doc_vals[row_i], prob_vals[row_i], raw_vals[row_i])
                            )
                        else:
                            helper_rows.append(("", "", ""))

                helper_start_row = 2
                for idx, (doc_v, prob_v, raw_v) in enumerate(helper_rows):
                    row_no = helper_start_row + idx
                    ws.cell(row=row_no, column=8, value=doc_v)  # H
                    ws.cell(row=row_no, column=9, value=prob_v)  # I
                    ws.cell(row=row_no, column=10, value=raw_v)  # J

                helper_end_row = helper_start_row + len(helper_rows) - 1
                for rank_i in range(1, max_rows_per_topic + 1):
                    out_row = 6 + rank_i
                    ws.cell(out_row, 1, f'=IF({rank_i}>$B$4,"",{rank_i})')
                    ws.cell(
                        out_row,
                        2,
                        (
                            f'=IFERROR(IF({rank_i}>$B$4,"",'
                            f'INDEX($H${helper_start_row}:$H${helper_end_row},($B$3-1)*$B$5+{rank_i})),"")'
                        ),
                    )
                    ws.cell(
                        out_row,
                        3,
                        (
                            f'=IFERROR(IF({rank_i}>$B$4,"",'
                            f'INDEX($I${helper_start_row}:$I${helper_end_row},($B$3-1)*$B$5+{rank_i})),"")'
                        ),
                    )
                    ws.cell(
                        out_row,
                        4,
                        (
                            f'=IFERROR(IF({rank_i}>$B$4,"",'
                            f'INDEX($J${helper_start_row}:$J${helper_end_row},($B$3-1)*$B$5+{rank_i})),"")'
                        ),
                    )

                topic_dv = DataValidation(
                    type="whole", operator="between", formula1="1", formula2=f"{k}"
                )
                topn_dv = DataValidation(
                    type="whole", operator="between", formula1="1", formula2="=$B$5"
                )
                ws.add_data_validation(topic_dv)
                ws.add_data_validation(topn_dv)
                topic_dv.add("B3")
                topn_dv.add("B4")

                ws.column_dimensions["A"].width = 10
                ws.column_dimensions["B"].width = 42
                ws.column_dimensions["C"].width = 14
                ws.column_dimensions["D"].width = 70
                ws.column_dimensions["H"].hidden = True
                ws.column_dimensions["I"].hidden = True
                ws.column_dimensions["J"].hidden = True

            with pd.ExcelWriter(word_topic_path, engine="openpyxl") as writer:
                beta_df.to_excel(writer, index=False, sheet_name="WordTopic")

                word_topk_sheet = "TopK_Results"
                ws_word = writer.book.create_sheet(title=word_topk_sheet)

                ws_word["A1"] = "WordTopic TopK 查詢"
                ws_word["A3"] = "topic number"
                ws_word["A4"] = "top number"
                ws_word["A5"] = "max rows/topic"
                ws_word["C3"] = '=IF(AND($B$3>=1,$B$3<=$E$1),"topic_"&$B$3,"")'

                ws_word["B3"] = 1
                max_rows_word_per_topic = (
                    int(beta_df.groupby("topic").size().max())
                    if not beta_df.empty
                    else 1
                )
                ws_word["B4"] = min(10, max_rows_word_per_topic)
                ws_word["B5"] = max_rows_word_per_topic
                ws_word["E1"] = int(k)

                ws_word["A6"] = "rank"
                ws_word["B6"] = "word"
                ws_word["C6"] = "probability"
                ws_word["D6"] = "log_probability"

                helper_word_rows = []
                for topic_i in range(1, k + 1):
                    topic_word_df = beta_df[beta_df["topic"] == topic_i].sort_values(
                        ["probability", "word"], ascending=[False, True]
                    )
                    word_vals = topic_word_df["word"].tolist()
                    prob_vals = topic_word_df["probability"].tolist()
                    log_prob_vals = topic_word_df["log_probability"].tolist()
                    for row_i in range(max_rows_word_per_topic):
                        if row_i < len(topic_word_df):
                            helper_word_rows.append(
                                (
                                    word_vals[row_i],
                                    prob_vals[row_i],
                                    log_prob_vals[row_i],
                                )
                            )
                        else:
                            helper_word_rows.append(("", "", ""))

                helper_word_start_row = 2
                for idx, (word_v, prob_v, log_prob_v) in enumerate(helper_word_rows):
                    row_no = helper_word_start_row + idx
                    ws_word.cell(row=row_no, column=8, value=word_v)  # H
                    ws_word.cell(row=row_no, column=9, value=prob_v)  # I
                    ws_word.cell(row=row_no, column=10, value=log_prob_v)  # J

                helper_word_end_row = helper_word_start_row + len(helper_word_rows) - 1
                for rank_i in range(1, max_rows_word_per_topic + 1):
                    out_row = 6 + rank_i
                    ws_word.cell(out_row, 1, f'=IF({rank_i}>$B$4,"",{rank_i})')
                    ws_word.cell(
                        out_row,
                        2,
                        (
                            f'=IFERROR(IF({rank_i}>$B$4,"",'
                            f'INDEX($H${helper_word_start_row}:$H${helper_word_end_row},($B$3-1)*$B$5+{rank_i})),"")'
                        ),
                    )
                    ws_word.cell(
                        out_row,
                        3,
                        (
                            f'=IFERROR(IF({rank_i}>$B$4,"",'
                            f'INDEX($I${helper_word_start_row}:$I${helper_word_end_row},($B$3-1)*$B$5+{rank_i})),"")'
                        ),
                    )
                    ws_word.cell(
                        out_row,
                        4,
                        (
                            f'=IFERROR(IF({rank_i}>$B$4,"",'
                            f'INDEX($J${helper_word_start_row}:$J${helper_word_end_row},($B$3-1)*$B$5+{rank_i})),"")'
                        ),
                    )

                topic_word_dv = DataValidation(
                    type="whole", operator="between", formula1="1", formula2=f"{k}"
                )
                topn_word_dv = DataValidation(
                    type="whole", operator="between", formula1="1", formula2="=$B$5"
                )
                ws_word.add_data_validation(topic_word_dv)
                ws_word.add_data_validation(topn_word_dv)
                topic_word_dv.add("B3")
                topn_word_dv.add("B4")

                ws_word.column_dimensions["A"].width = 10
                ws_word.column_dimensions["B"].width = 20
                ws_word.column_dimensions["C"].width = 14
                ws_word.column_dimensions["D"].width = 18
                ws_word.column_dimensions["H"].hidden = True
                ws_word.column_dimensions["I"].hidden = True
                ws_word.column_dimensions["J"].hidden = True

            with pd.ExcelWriter(topic_doc_path) as writer:
                raw_pivot.to_excel(writer, index=False, sheet_name="TopicDoc")
                raw_long.to_excel(writer, index=False, sheet_name="TopicDoc_Long")

            on_append(
                f"LDA 適配完成。\n"
                f"  TopicPart → {topic_part_path}\n"
                f"  WordTopic → {word_topic_path}\n"
                f"  TopicDoc  → {topic_doc_path}"
            )
            on_done()

        except Exception as e:
            on_error(str(e))

    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    return t
